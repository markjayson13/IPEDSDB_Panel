"""
Integration tests for Stage 09 panel-dictionary generation.

Focus:
- schema-driven row order
- dictionary text merge
- controlled YEAR metadata when not present in dictionary_lake
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from helpers import run_script, write_parquet


def test_build_panel_dictionary_merges_schema_and_dictionary_text(tmp_path: Path) -> None:
    root = tmp_path / "data_root"
    input_path = root / "Panels" / "panel_clean_analysis_2022_2023.parquet"
    dictionary_path = root / "Dictionary" / "dictionary_lake.parquet"
    output_path = root / "Panels" / "panel_dictionary.csv"

    write_parquet(
        input_path,
        [
            {"year": 2023, "UNITID": 100663, "INSTNM": "Example B", "CONTROL": 2},
        ],
    )
    write_parquet(
        dictionary_path,
        [
            {
                "varname": "INSTNM",
                "varTitle": "Institution name",
                "longDescription": "Institution name used in reporting and public lookup outputs.",
                "DataType": "char",
            },
            {
                "varname": "CONTROL",
                "varTitle": "Control",
                "longDescription": "Institutional control sector.",
                "DataType": "disc",
            },
        ],
    )

    result = run_script(
        "Scripts/09_build_panel_dictionary.py",
        "--input",
        input_path,
        "--dictionary",
        dictionary_path,
        "--output",
        output_path,
    )

    assert result.returncode == 0, result.stdout
    out = pd.read_csv(output_path)
    assert out["varname"].tolist() == ["YEAR", "UNITID", "INSTNM", "CONTROL"]

    year_row = out[out["varname"] == "YEAR"].iloc[0]
    assert year_row["varTitle"] == "IPEDS reporting year"
    assert "institution-year panel" in year_row["longDescription"]

    instnm_row = out[out["varname"] == "INSTNM"].iloc[0]
    assert instnm_row["varTitle"] == "Institution name"
    assert instnm_row["dictionaryDataType"] == "char"

    control_row = out[out["varname"] == "CONTROL"].iloc[0]
    assert control_row["panelDataType"] in {"int64", "int64[pyarrow]", "int32", "int32[pyarrow]"}


def test_build_panel_dictionary_writes_formatted_excel_workbook(tmp_path: Path) -> None:
    root = tmp_path / "data_root"
    input_path = root / "Panels" / "panel_clean_analysis_2022_2023.parquet"
    dictionary_path = root / "Dictionary" / "dictionary_lake.parquet"
    output_path = root / "Panels" / "panel_dictionary.xlsx"

    write_parquet(
        input_path,
        [
            {"year": 2023, "UNITID": 100663, "INSTNM": "Example B", "CONTROL": 2},
        ],
    )
    write_parquet(
        dictionary_path,
        [
            {
                "varname": "INSTNM",
                "varTitle": "Institution name",
                "longDescription": "Institution name used in reporting and public lookup outputs.",
                "DataType": "char",
            },
            {
                "varname": "CONTROL",
                "varTitle": "Control",
                "longDescription": "Institutional control sector.",
                "DataType": "disc",
            },
        ],
    )

    result = run_script(
        "Scripts/09_build_panel_dictionary.py",
        "--input",
        input_path,
        "--dictionary",
        dictionary_path,
        "--output",
        output_path,
    )

    assert result.returncode == 0, result.stdout
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["panel_dictionary", "about"]

    ws = wb["panel_dictionary"]
    assert ws.freeze_panes == "A2"
    headers = [cell.value for cell in ws[1]]
    assert headers == ["column_order", "varname", "varTitle", "longDescription", "panelDataType", "dictionaryDataType"]
    assert ws["B2"].value == "YEAR"
    assert ws["C2"].value == "IPEDS reporting year"

    about = wb["about"]
    assert about["A1"].value == "Panel dictionary export"
    assert str(input_path) == about["B6"].value
