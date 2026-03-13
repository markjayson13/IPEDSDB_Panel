#!/usr/bin/env python3
"""
Stage 09: build a panel-specific variable dictionary for a stitched wide panel.

Reads:
- a stitched wide or cleaned wide parquet panel
- `Dictionary/dictionary_lake.parquet`

Writes:
- a panel-level dictionary keyed to the actual output columns
- `.csv` for plain export or `.xlsx` for a formatted workbook
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, help="Input stitched wide parquet")
    p.add_argument("--dictionary", required=True, help="dictionary_lake.parquet")
    p.add_argument("--output", required=True, help="Output .csv or .xlsx path")
    return p.parse_args()


def best_text(series: pd.Series) -> str:
    vals = [str(v).strip() for v in series.dropna().tolist()]
    vals = [v for v in vals if v and v.lower() not in {"nan", "none", "<na>", "nat"}]
    if not vals:
        return ""
    vals = sorted(set(vals), key=lambda x: (-len(x), x))
    return vals[0]


def summarize_dictionary(path: Path) -> pd.DataFrame:
    cols = ["varname", "varTitle", "longDescription", "DataType"]
    df = pd.read_parquet(path, columns=cols)
    for col in cols:
        if col not in df.columns:
            df[col] = ""
    df["varname"] = df["varname"].fillna("").astype(str).str.upper().str.strip()
    df = df[df["varname"] != ""].copy()
    out = (
        df.groupby("varname", as_index=False)
        .agg(
            varTitle=("varTitle", best_text),
            longDescription=("longDescription", best_text),
            dictionaryDataType=("DataType", best_text),
        )
    )
    return out


def panel_schema_df(path: Path) -> pd.DataFrame:
    schema = pq.read_schema(path)
    rows = []
    for idx, name in enumerate(schema.names):
        field = schema.field(name)
        rows.append(
            {
                "column_order": idx,
                "varname": str(name).upper().strip(),
                "panelDataType": str(field.type),
            }
        )
    return pd.DataFrame(rows)


def build_reference_df(input_path: Path, dictionary_path: Path) -> pd.DataFrame:
    schema_df = panel_schema_df(input_path)
    dict_df = summarize_dictionary(dictionary_path)
    ref = schema_df.merge(dict_df, on="varname", how="left")

    # Controlled metadata for stitched-wide panel keys.
    ref.loc[ref["varname"] == "YEAR", "varTitle"] = (
        ref.loc[ref["varname"] == "YEAR", "varTitle"].fillna("").replace("", "IPEDS reporting year")
    )
    ref.loc[ref["varname"] == "YEAR", "longDescription"] = (
        ref.loc[ref["varname"] == "YEAR", "longDescription"]
        .fillna("")
        .replace("", "IPEDS reporting year carried by the stitched institution-year panel.")
    )
    ref.loc[ref["varname"] == "UNITID", "varTitle"] = (
        ref.loc[ref["varname"] == "UNITID", "varTitle"].fillna("").replace("", "Institution identifier (panel key)")
    )
    ref.loc[ref["varname"] == "UNITID", "longDescription"] = (
        ref.loc[ref["varname"] == "UNITID", "longDescription"]
        .fillna("")
        .replace("", "Stable institution identifier used as the unit key in the stitched institution-year panel.")
    )

    for col in ["varTitle", "longDescription", "dictionaryDataType"]:
        ref[col] = ref[col].fillna("").astype(str)

    return ref[["column_order", "varname", "varTitle", "longDescription", "panelDataType", "dictionaryDataType"]]


def write_excel(ref: pd.DataFrame, input_path: Path, dictionary_path: Path, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "panel_dictionary"
    headers = list(ref.columns)
    ws.append(headers)
    for row in ref.itertuples(index=False, name=None):
        ws.append(list(row))

    header_fill = PatternFill(fill_type="solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    subhead_fill = PatternFill(fill_type="solid", fgColor="E9F0F7")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24

    width_map = {
        "A": 14,
        "B": 18,
        "C": 42,
        "D": 110,
        "E": 18,
        "F": 20,
    }
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width

    wrap_cols = {3, 4, 5, 6}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="PanelDictionary", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    about = wb.create_sheet("about")
    about["A1"] = "Panel dictionary export"
    about["A1"].font = Font(bold=True, size=14)
    about["A3"] = "What this workbook is"
    about["A3"].font = Font(bold=True)
    about["A3"].fill = subhead_fill
    about["A4"] = (
        "A formatted dictionary for the actual columns present in the panel output. "
        "Open the 'panel_dictionary' sheet for the variable list."
    )
    about["A6"] = "Source panel"
    about["A6"].font = Font(bold=True)
    about["B6"] = str(input_path)
    about["A7"] = "Source metadata"
    about["A7"].font = Font(bold=True)
    about["B7"] = str(dictionary_path)
    about["A8"] = "Rows in dictionary"
    about["A8"].font = Font(bold=True)
    about["B8"] = int(len(ref))
    about["A10"] = "Columns in main sheet"
    about["A10"].font = Font(bold=True)
    about["A11"] = "column_order"
    about["A12"] = "varname"
    about["A13"] = "varTitle"
    about["A14"] = "longDescription"
    about["A15"] = "panelDataType"
    about["A16"] = "dictionaryDataType"
    about.column_dimensions["A"].width = 26
    about.column_dimensions["B"].width = 120
    for row in about.iter_rows(min_row=1, max_row=16, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    dictionary_path = Path(args.dictionary)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    ref = build_reference_df(input_path, dictionary_path)
    if out_path.suffix.lower() == ".xlsx":
        write_excel(ref, input_path, dictionary_path, out_path)
    else:
        ref.to_csv(out_path, index=False)
    print(f"Wrote {len(ref):,} rows to {out_path}")


if __name__ == "__main__":
    main()
